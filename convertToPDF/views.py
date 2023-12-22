from django.shortcuts import render
from docx2pdf import convert
from django.core.files.storage import FileSystemStorage
from PIL import Image
import os
import win32com.client
import pythoncom 
import pdfkit
from django.http import HttpResponse
from openpyxl import load_workbook
from reportlab.pdfgen import canvas



def wordToPdf(request):
    if request.method == 'POST':
        select_file = request.FILES['select_file']  
        if select_file:
            file_name = select_file.name
            file_url = None
            fss = FileSystemStorage()
            file = fss.save(file_name, select_file)
            file_url = fss.url(file)
            name, extension = os.path.splitext(file_name) 
            docx_file = r"C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}".format(file_name)
            pdf_file = f"C:\\Users\\kushwah\\OneDrive\Desktop\\convert\\pdfConverter\\media\\{name}.pdf"
            print("word to pdf")
            try:
                pythoncom.CoInitialize()
                word = win32com.client.Dispatch('Word.Application')
                doc = word.Documents.Open(docx_file)
                doc.SaveAs(pdf_file, FileFormat=17)  # 17 represents PDF file format
                doc.Close()
                word.Quit()
                pythoncom.CoUninitialize()
            except Exception as e:
                print(f"Error: {str(e)}")
                return HttpResponse("Error converting Word to PDF.") 
            # Provide a link to download the converted PDF
            context = {'file_url': docx_file, 'pdf_file_url': fss.url(f"{name}.pdf")}
            print(context)
            return render(request, 'convertToPdf/wordToPdf.html', context)
        else:
            return HttpResponse("No file uploaded.")
    return render(request, 'convertToPdf/wordToPdf.html')

def excelToPdf(request):
    if request.method == 'POST':    
        file_name = request.FILES['select_file']
        fss = FileSystemStorage()
        file = fss.save(file_name.name, file_name)
        file_url = fss.url(file)
        file_name = str(file_name)
        excel_file = r'C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}'.format(file_name)
        name, extension = os.path.splitext(file_name)
        pdf_file = r"C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}.pdf".format(name)
        print("excel_file--",excel_file)
        print("pdf_file--",pdf_file)
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            # Create a PDF file
            c = canvas.Canvas(pdf_file)
            # Iterate through the Excel sheet and write to the PDF
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    c.drawString(100, 800, str(cell))
                c.showPage()
            # Close the PDF file
                c.save()
        except Exception as e:
            print(f"Error: {str(e)}")
        context = {'file_url': excel_file, 'pdf_file_url': fss.url(f"{name}.pdf")}
        print(context)
        return render(request, 'convertToPdf/excelToPdf.html', context)    
        # return render(request,'convertToPdf/excelToPdf.html',{'file_url': file_url})
    return render(request,'convertToPdf/excelToPdf.html')


def htmlToPdf(request):
    if request.method == 'POST':
        url = request.POST.get('url')
        output_pdf = r'C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\ou.pdf'
        print("url-",url)
        if url:
            try:
                pdfkit.from_url(url, output_pdf)
                print(f"PDF successfully generated at {pdf_file}")
                return HttpResponse(f"PDF successfully generated at {pdf_file}")
            except Exception as e:
                print(f"Error: {str(e)}")
                return HttpResponse(f"Error: {str(e)}")
        return render(request, 'convertToPdf/htmlToPdf.html')
    return render(request, 'convertToPdf/htmlToPdf.html')


def jpgToPdf(request):
    if request.method == 'POST':    
        img=request.FILES['select_file']
        fss = FileSystemStorage()
        file = fss.save(img.name, img)
        file_url = fss.url(file)
        img = str(img)
        image_1 = Image.open(r'C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\%s' % img)
        im_1 = image_1.convert('RGB')
        name, extension = os.path.splitext(img)
        im_1.save(r'C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}.pdf'.format(name))
        context = {'file_url': image_1, 'pdf_file_url': fss.url(f"{name}.pdf")}
        print(context)
        return render(request, 'convertToPdf/jpdToPdf.html', context)  
    return render(request,'convertToPdf/jpdToPdf.html')

def powerpointToPdf(request):
    if request.method == 'POST':
        if 'select_file' in request.FILES: 
            file_name = request.FILES['select_file']
            fss = FileSystemStorage()
            file = fss.save(file_name.name, file_name)
            file_url = fss.url(file)
            file_name = str(file_name)
            ppt_file = r'C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}'.format(file_name)
            print(ppt_file)
            name, extension = os.path.splitext(file_name)
            pdf_file = r"C:\Users\kushwah\OneDrive\Desktop\convert\pdfConverter\media\{}.pdf".format(name)
            print("PowerPoint To PDF")
            try:
                pythoncom.CoInitialize()
                powerpoint = win32com.client.Dispatch("PowerPoint.Application")
                presentation = powerpoint.Presentations.Open(ppt_file)
                presentation.SaveAs(pdf_file, 32)  # 32 represents PDF file format
                presentation.Close()
                powerpoint.Quit()
                pythoncom.CoUninitialize()
            except Exception as e:
                print(f"An error occurred: {str(e)}")
            #  Provide a link to download the converted PDF
            context = {'file_url': ppt_file, 'pdf_file_url': fss.url(f"{name}.pdf")}
            print(context)    
        return render(request,'convertToPdf/powerponitToPdf.html',context)    
    return render(request,'convertToPdf/powerponitToPdf.html')
